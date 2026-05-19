"""WIPS '발명의 설명' 본문을 상세보기 URL에서 긁어와 JSON으로 누적 저장.

사용:
    python3 scripts/scrape_descriptions.py            # 전체 729건
    python3 scripts/scrape_descriptions.py --limit 5  # 처음 5건만
    python3 scripts/scrape_descriptions.py --limit 5 --force  # 이미 있어도 재수집
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "list" / "datas_list.xlsx"
OUT = ROOT / "data" / "descriptions_scraped.json"

DETAIL_COL = "상세보기 링크(비로그인)"
CTRY_COL = "국가코드"
APPNO_COL = "출원번호"
PUBNO_COL = "공개번호"
REGNO_COL = "등록번호"
TITLE_COL = "발명의 명칭"

POPUP_URL = "https://sd.wips.co.kr/wipslink/doc/docLargeDataPopup.wips"
HTML_URL_TMPL = "https://sd.wips.co.kr/wipslink/doc/docLargePopupLoading.wips?skey={skey}&langCd={lang}"

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"


def load_rows():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {k: row[idx[k]] for k in (DETAIL_COL, CTRY_COL, APPNO_COL, PUBNO_COL, REGNO_COL, TITLE_COL)}


def skey_from_url(url: str) -> str | None:
    if not url:
        return None
    q = parse_qs(urlparse(url).query)
    v = q.get("skey", [None])[0]
    return v


def fetch_description(session: requests.Session, skey: str, lang: str) -> dict:
    """Return {'html': ..., 'text': ..., 'status': int}."""
    # The popup uses a small bootstrap page that POSTs to docLargeDataPopup.wips
    # We just POST directly.
    r = session.post(
        POPUP_URL,
        data={"skey": skey, "langCd": lang, "nmtLangCd": ""},
        headers={
            "User-Agent": UA,
            "Referer": HTML_URL_TMPL.format(skey=skey, lang=lang),
            "Origin": "https://sd.wips.co.kr",
        },
        timeout=60,
    )
    return {"status": r.status_code, "html": r.text}


def extract_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    # The description body sits in #devDescriptionView or as the main content
    container = soup.select_one("#devDescriptionView") or soup.select_one("body") or soup
    # Drop scripts/styles
    for tag in container.find_all(["script", "style"]):
        tag.decompose()
    text = container.get_text("\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def load_existing() -> dict:
    if OUT.exists():
        return json.loads(OUT.read_text(encoding="utf-8"))
    return {}


def save(store: dict):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="0 = no limit")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--sleep", type=float, default=0.8)
    args = ap.parse_args()

    store = load_existing()
    session = requests.Session()

    count = 0
    for row in load_rows():
        if args.limit and count >= args.limit:
            break
        url = row[DETAIL_COL]
        skey = skey_from_url(url)
        if not skey:
            continue
        ctry = (row[CTRY_COL] or "").strip()
        doc_id = row[PUBNO_COL] or row[REGNO_COL] or row[APPNO_COL] or skey
        if isinstance(doc_id, str):
            doc_id = doc_id.strip()
        # skey를 항상 포함시켜 doc_id가 비어 있어도 key가 유일해지도록 함
        key = f"{ctry}_{doc_id}_{skey}" if not doc_id or doc_id == skey else f"{ctry}_{doc_id}"

        if key in store and not args.force:
            print(f"[skip] {key} (이미 수집됨)")
            count += 1
            continue

        print(f"[fetch] {key}  skey={skey}  lang={ctry}")
        try:
            res = fetch_description(session, skey, ctry)
        except Exception as e:
            print(f"  ! 에러: {e}")
            continue

        if res["status"] != 200:
            print(f"  ! HTTP {res['status']}")
            continue

        text = extract_text(res["html"])
        n = len(text)
        print(f"  ok  {n} chars")

        store[key] = {
            "skey": skey,
            "ctry": ctry,
            "doc_id": doc_id,
            "title": row[TITLE_COL],
            "url": url,
            "lang_cd": ctry,
            "char_len": n,
            "text": text,
        }
        save(store)
        count += 1
        time.sleep(args.sleep)

    print(f"\n완료: {count}건 처리, 누적 {len(store)}건 저장 → {OUT}")


if __name__ == "__main__":
    sys.exit(main())
